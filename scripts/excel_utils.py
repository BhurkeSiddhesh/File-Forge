"""
Excel/CSV utilities for File Forge.

Pure-Python implementations using openpyxl + pandas + reportlab.
Excel→PDF renders sheets as styled tables, NOT as a Microsoft-Office-grade rendering;
cell colors, merged cells, charts, and conditional formatting are approximated or dropped.
"""
import csv
import uuid
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
)

from scripts.utils import branded_filename, original_stem


def _sanitize_cell(value) -> str:
    if value is None:
        return ""
    return str(value)


def excel_to_pdf(input_path: str, output_dir: str) -> str:
    """Render every sheet of an .xlsx workbook as a table in a PDF."""
    input_file = Path(input_path)
    output_file = Path(output_dir) / branded_filename(input_file, "pdf")

    wb = load_workbook(input_file, data_only=True, read_only=True)
    try:
        if not wb.sheetnames:
            raise ValueError("Workbook has no sheets.")

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "SheetTitle", parent=styles["Heading2"], spaceAfter=10
        )

        doc = SimpleDocTemplate(
            str(output_file),
            pagesize=landscape(A4),
            leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24,
        )
        story = []

        for s_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            story.append(Paragraph(f"Sheet: {sheet_name}", title_style))

            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([_sanitize_cell(c) for c in row])

            if not rows:
                story.append(Paragraph("<i>(empty sheet)</i>", styles["BodyText"]))
            else:
                # Cap dimensions so the PDF doesn't explode for huge sheets.
                max_rows = 500
                max_cols = 20
                truncated_rows = rows[:max_rows]
                truncated_rows = [r[:max_cols] for r in truncated_rows]

                table = Table(truncated_rows, repeatRows=1)
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]))
                story.append(table)

                if len(rows) > max_rows or any(len(r) > max_cols for r in rows):
                    story.append(Spacer(1, 6))
                    story.append(Paragraph(
                        f"<i>Truncated to first {max_rows} rows × {max_cols} columns.</i>",
                        styles["BodyText"],
                    ))

            if s_idx < len(wb.sheetnames) - 1:
                story.append(PageBreak())

        doc.build(story)
    finally:
        wb.close()

    return str(output_file)


def csv_to_xlsx(input_path: str, output_dir: str, delimiter: str = ",") -> str:
    """Convert a CSV file into a single-sheet .xlsx workbook."""
    input_file = Path(input_path)
    output_file = Path(output_dir) / branded_filename(input_file, "xlsx")

    delimiter = delimiter or ","
    # Normalize the literal "\t" escape to a real tab BEFORE validating length,
    # so a real two-char input like "||" is rejected with a clear ValueError
    # instead of falling through to csv.reader and raising a raw TypeError.
    if delimiter == "\\t":
        delimiter = "\t"
    if len(delimiter) != 1:
        raise ValueError("delimiter must be a single character (use ',' '\\t' ';' '|').")

    wb = Workbook()
    ws = wb.active
    ws.title = original_stem(input_file)[:31] or "Sheet1"

    with input_file.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            ws.append(row)

    wb.save(output_file)
    return str(output_file)


def xlsx_to_csv(input_path: str, output_dir: str, sheet: str = None) -> str:
    """Convert an .xlsx sheet to CSV. If `sheet` is None, uses the first sheet."""
    input_file = Path(input_path)

    wb = load_workbook(input_file, data_only=True, read_only=True)
    try:
        if sheet:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet]
            sheet_label = sheet
        else:
            ws = wb[wb.sheetnames[0]]
            sheet_label = wb.sheetnames[0]

        # Sanitize sheet name for filename use.
        safe_label = "".join(c for c in sheet_label if c.isalnum() or c in "-_") or "sheet"
        output_file = Path(output_dir) / branded_filename(input_file, "csv")

        with output_file.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([_sanitize_cell(c) for c in row])
    finally:
        wb.close()

    return str(output_file)


def merge_excel_files(input_paths: List[str], output_dir: str) -> str:
    """Merge multiple .xlsx files into one workbook (each input sheet → one output sheet)."""
    if not input_paths:
        raise ValueError("No input files provided for merging.")
    if len(input_paths) < 2:
        raise ValueError("Provide at least two Excel files to merge.")

    output_file = Path(output_dir) / f"merged_{uuid.uuid4().hex[:8]}.xlsx"
    out_wb = Workbook()
    # Remove the default empty sheet so we start clean.
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    used_names = set()

    def unique_name(base: str) -> str:
        base = (base or "Sheet")[:25]  # leave room for suffix
        candidate = base
        i = 1
        while candidate in used_names or len(candidate) > 31:
            i += 1
            candidate = f"{base}_{i}"[:31]
        used_names.add(candidate)
        return candidate

    for path in input_paths:
        src = Path(path)
        wb = load_workbook(src, data_only=True, read_only=True)
        try:
            for sheet_name in wb.sheetnames:
                src_ws = wb[sheet_name]
                target_name = unique_name(f"{original_stem(src)}_{sheet_name}")
                dst_ws = out_wb.create_sheet(target_name)
                for row in src_ws.iter_rows(values_only=True):
                    dst_ws.append(list(row))
        finally:
            wb.close()

    if not out_wb.sheetnames:
        # Edge case: every input was empty.
        out_wb.create_sheet("Sheet1")

    out_wb.save(output_file)
    return str(output_file)
